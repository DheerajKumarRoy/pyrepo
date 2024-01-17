"""
This is a Expense Tracking app built with Python Kivy.

Author: Dheeraj k. Roy
Date: 12/05/2023
"""
from kivy.app import App
from kivy.uix.gridlayout import GridLayout
from kivy.uix.boxlayout import BoxLayout
from kivy.uix.button import Button
from kivy.uix.label import Label
from kivy.uix.image import Image
from kivy.uix.textinput import TextInput
from kivy.uix.scrollview import ScrollView
from kivy.uix.checkbox import CheckBox
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font as font, Alignment, Border, Side
import pandas as pd
import os

# directories handling
parent_path = 'ExTrack'

class MyGrid(GridLayout):
    def __init__(self, **kwargs):
        super(MyGrid, self).__init__(**kwargs)
        self.cols = 1  # Set the number of columns in the GridLayout
        # inner window
        self.window = GridLayout()
        self.window.cols = 2
        self.window.size_hint = (1, 0.7)
        # check box
        self.cb_label = BoxLayout()
        self.checkbox_grid = BoxLayout()
        # add button
        self.b1 = GridLayout()
        self.b1.cols = 3
        # export button
        self.exp = BoxLayout()
        # Entries
        self.entries = BoxLayout()
        # bottom
        self.view = GridLayout()
        self.view.cols = 1
        self.view.size_hint_y = 7
        # logo
        self.add_widget(Image(source='logo/logo.png', size_hint=(1, 2)))
        # Title_label
        self.item = Label(
            text='Expense Tracker',
            bold=True,
            # font_size=40,
            color='#00FFCE'
        )
        self.add_widget(self.item)
        # user input
        self.item = TextInput(
            hint_text='Item?',
            multiline=False,
            # padding_y=(20,20),
            size_hint=(0.5, 0.5),
            # color='#00FFCE'
        )
        self.item.bind(text=self.on_text)
        self.window.add_widget(self.item)
        self.Price = TextInput(
            hint_text='Price?',
            multiline=False,
            # padding_y=(20,20),
            size_hint=(0.5, 0.5),
            # color='#00FFCE'
        )
        self.Price.bind(text=self.on_text)
        self.window.add_widget(self.Price)
        self.add_widget(self.window)

        # add button
        self.add = Button(text='Add',
                          size_hint=(0.5, 0.25),
                          # font_size=20,
                          bold=True,
                          disabled=True
                          )
        self.add.bind(on_press=self.callback)
        self.lab1 = Label(text="")
        self.b1.add_widget(self.lab1)
        self.b1.add_widget(self.add)
        self.lab2 = Label(text="")
        self.b1.add_widget(self.lab2)
        self.add_widget(self.b1)
        # checkboxses
        self.Text_file_label = Label(text="Text")
        self.XL_file_label = Label(text="Excel")
        self.Reset_label = Label(text='Reset')
        self.Delete_label = Label(text='Delete')

        self.cb_label.add_widget(self.Reset_label)
        self.cb_label.add_widget(self.Delete_label)
        self.cb_label.add_widget(self.Text_file_label)
        self.cb_label.add_widget(self.XL_file_label)
        self.add_widget(self.cb_label)

        self.checkbox_reset = CheckBox(active=False)
        self.checkbox_delete = CheckBox(active=False)
        self.checkbox_txt = CheckBox(active=False)
        self.checkbox_xl = CheckBox(active=False)

        self.checkbox_grid.add_widget(self.checkbox_reset)
        self.checkbox_grid.add_widget(self.checkbox_delete)
        # self.checkbox_grid.add_widget(self.resetlab4)
        self.checkbox_grid.add_widget(self.checkbox_txt)
        self.checkbox_grid.add_widget(self.checkbox_xl)
        self.add_widget(self.checkbox_grid)

        # save button
        self.resetlab1 = Label(text="")
        self.resetlab2 = Label(text="")
        self.resetlab3 = Label(text="")
        self.resetlab4 = Label(text="")
        self.save = Button(text='save',
                           disabled=True,
                           # font_size=14,
                           )
        self.save.bind(on_press=self.Export)

        # reset button
        self.reset = Button(text="reset",
                            disabled=True,
                            # font_size=14
                            )
        self.reset.bind(on_press=self.Reset)

        self.exp.add_widget(self.resetlab1)
        self.exp.add_widget(self.reset)
        self.exp.add_widget(self.resetlab2)
        self.exp.add_widget(self.resetlab4)
        self.exp.add_widget(self.save)
        self.exp.add_widget(self.resetlab3)
        self.add_widget(self.exp)

        # Show data button
        self.showdata = Button(text='Entries',
                               disabled=True
                               # font_size=14,
                               )
        self.showdata.bind(on_release=self.info)

        self.lab5 = Label(text="")
        self.entries.add_widget(self.lab5)
        self.lab6 = Label(text="")
        self.entries.add_widget(self.lab6)
        self.entries.add_widget(self.showdata)
        self.lab7 = Label(text="")
        self.entries.add_widget(self.lab7)
        self.lab8 = Label(text="")
        self.entries.add_widget(self.lab8)
        self.add_widget(self.entries)

        if os.path.isfile(parent_path+'/App_Data/'+".Expense_Details.xlsx"):
            self.reset.disabled = False
            self.save.disabled = False
            self.showdata.disabled = False

        # showdata
        self.scroll_view = ScrollView()
        self.label = Label(size_hint=(None, None),
                           text_size=(None, None),
                           padding=(20, 20),
                           halign='left',
                           valign='top',
                           markup=True,
                           # font_size=20
                           )
        self.label.bind(texture_size=self.label.setter('size'))
        self.scroll_view.add_widget(self.label)
        self.view.add_widget(self.scroll_view)
        self.add_widget(self.view)

    def callback(self, instance):
        self.time = datetime.now().strftime('%d/%m/%y-%H:%M')
        self.xltime = datetime.now().strftime('%d-%m-%Y')
        self.a, self.b = self.item.text, self.Price.text
        self.b = float(self.b)
        self.remarks = ''
        with open(parent_path+'/App_Data'+'/.data', 'a') as f:
            f.write(
                f'\n({self.time})-[{self.item.text}---{self.Price.text}Rs.]')

            # clear user input after submit
        self.item.text = ''
        self.Price.text = ''
        with open(parent_path+'/App_Data'+'/.data') as f:
            S_data = f.read()
            self.label.text = S_data
        self.reset.disabled = False
        self.save.disabled = False
        self.showdata.disabled = False

        # Excel
        wbName = f".Expense_Details.xlsx"
        # if file is not present
        if not os.path.isfile(parent_path+'/App_Data/'+wbName):
            wsName = "Expenses"
            wb = Workbook()
            ws = wb.active
            ws.title = wsName

            ws['A2'] = 'Date'
            ws['B2'] = 'Item'
            ws['C2'] = 'Cost(Rs.)'
            ws['D2'] = 'Remarks'

            self.remarks = ''
            if self.b < 200:
                self.remarks = 'SmallAmmount'
            elif 200 < self.b < 2000:
                self.remarks = 'NormalAmmount'
            elif 2000 < self.b < 5000:
                self.remarks = 'MediumAmmount'
            elif 5000 < self.b:
                self.remarks = 'HighAmmount'
            row = [self.xltime, self.a, self.b, self.remarks]
            ws.append(row)
            values = [cell.value for cell in ws['C']]
            tItem = len(values)
            self.sum = sum(values[2:])

            # total Number of Days since fist entry
            firstCell = ws['A3']
            firstDate = firstCell.value
            currentDate_par = datetime.now().strptime(self.xltime, "%d-%m-%Y")
            firstDatePar = datetime.now().strptime(firstDate, "%d-%m-%Y")
            delta = currentDate_par-firstDatePar
            numDays = delta.days
            # DETAILS
            ws['A1'] = f'Total_Days={numDays}'
            ws['B1'] = f'Total_Payments={tItem-2}'
            ws['C1'] = f'Total_Spend={self.sum}Rs.'
            ws['D1'] = '-'

            # bold headings
            ft = font(bold=True)
            for row in ws['A1:D2']:
                for cell in row:
                    cell.font = ft

            wb.save(parent_path+'/App_Data/'+wbName)
        else:
            wb = load_workbook(parent_path+'/App_Data/'+wbName)
            ws = wb.active
            last_row = ws.max_row
            self.remarks = ''
            if self.b < 200:
                self.remarks = 'SmallAmmount'
            elif 200 < self.b < 2000:
                self.remarks = 'NormalAmmount'
            elif 2000 < self.b < 5000:
                self.remarks = 'MediumAmmount'
            elif 5000 < self.b:
                self.remarks = 'HighAmmount'

            row = [self.xltime, self.a, self.b, self.remarks]
            ws.append(row)
            values = [cell.value for cell in ws['C']]
            tItem = len(values)
            self.sum = sum(values[2:])

            # total Number of Days since fist entry
            firstCell = ws['A3']
            firstDate = firstCell.value
            currentDate_par = datetime.now().strptime(self.xltime, "%d-%m-%Y")
            firstDatePar = datetime.now().strptime(firstDate, "%d-%m-%Y")
            delta = currentDate_par-firstDatePar
            numDays = delta.days
            # DETAILS
            ws['A1'] = f'Total_Days={numDays}'
            ws['B1'] = f'Total_Payments={tItem-2}'
            ws['C1'] = f'Total_Spend={self.sum}Rs.'
            ws['D1'] = '-'
            # bold headings
            ft = font(bold=True)
            for row in ws['A1:D2']:
                for cell in row:
                    cell.font = ft
            # Create an Alignment object and set the horizontal and vertical alignment to center
            alignment = Alignment(horizontal='center', vertical='center')
            # Apply the alignment to the cell
            for row in ws:
                for cell in row:
                    cell.alignment = alignment
            # auto adjusting columns
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                # Iterate over each cell in the column and find the maximum length of content
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except TypeError:
                        pass
                # Set the adjusted width of the column based on the maximum length of content
                # Adding extra width for padding
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column_letter].width = adjusted_width
            # Create a Border object with desired border style
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            # Apply the border to the range of cells
            for row in ws:
                for cell in row:
                    cell.border = border
            wb.save(parent_path+'/App_Data/'+wbName)

    def Export(self, instance):
        if os.path.exists(parent_path+'/App_Data/'+".Expense_Details.xlsx"):
            if self.checkbox_txt.active:
                CurrentTime = datetime.now().strftime('%d%B%Y_%I%M%S%p')
                f_name = parent_path+f"/Saved/My_Data_{CurrentTime}.txt"
                df = pd.read_excel(parent_path+'/App_Data/' +
                                   ".Expense_Details.xlsx")
                df = str(df)
                with open(f_name, 'w') as f:
                    f.write(df)
                self.label.text = f"Entries saved in a text file at '{parent_path}/Saved'"
                if self.checkbox_xl.active:
                    CurrentTime = datetime.now().strftime('%d%B%Y_%H_%M_%S')
                    xl_name = parent_path + \
                        f'/Saved/Expense_Details_{CurrentTime}.xlsx'
                    wb = load_workbook(
                        parent_path+'/App_Data/'+".Expense_Details.xlsx")
                    ws = wb.active
                    wb.save(xl_name)
                    self.label.text = f"Entries saved in text & excel file at '{parent_path}/Saved'!"
                self.save.disabled = True
            elif self.checkbox_xl.active:
                CurrentTime = datetime.now().strftime('%d%B%Y_%H_%M_%S')
                xl_name = parent_path + \
                    f'/Saved/Expense_Details_{CurrentTime}.xlsx'
                wb = load_workbook(parent_path+'/App_Data/' +
                                   ".Expense_Details.xlsx")
                ws = wb.active
                wb.save(xl_name)
                self.label.text = f"Entries saved in a excel file at '{parent_path}/Saved'!"
                self.save.disabled = True
            else:
                self.label.text = "Please check Text/excel box First!"
        else:
            self.label.text = "Please Add Data First!"
            self.save.disabled = True

    def Reset(self, instance):
        if os.path.exists(parent_path+'/App_Data/'+".Expense_Details.xlsx"):
            if self.checkbox_reset.active:
                CurrentTime = datetime.now().strftime('%d%B%Y_%I%M%S%p')
                f_name = parent_path + \
                    f"/Reset_Backups/Reset_Backup_{CurrentTime}.txt"
                df = pd.read_excel(parent_path+'/App_Data/' +
                                   ".Expense_Details.xlsx")
                df = str(df)
                with open(f_name, 'w') as f:
                    f.write(df)
                CurrentTime = datetime.now().strftime('%d%B%Y_%H_%M_%S')
                xl_name = parent_path + \
                    f'/Reset_Backups/Reset_Backup{CurrentTime}.xlsx'
                wb = load_workbook(parent_path+'/App_Data/' +
                                   ".Expense_Details.xlsx")
                ws = wb.active
                wb.save(xl_name)
                self.label.text = f"Entries Reset & Backup copy saved at '{parent_path}/Reset_Backups'!"
                self.reset.disabled = True
                self.save.disabled = True
                os.remove(parent_path+'/App_Data/'+".Expense_Details.xlsx")
                os.remove(parent_path+'/App_Data/'+'.data')
            elif self.checkbox_delete.active:
                os.remove(parent_path+'/App_Data/'+".Expense_Details.xlsx")
                os.remove(parent_path+'/App_Data/'+'.data')
                self.label.text = "All Entries Deleted!"
                self.save.disabled = True
                self.reset.disabled = True
            else:
                self.label.text = "Please check Reset/Delete box First!"
        else:
            self.reset.disabled = True
            self.save.disabled = True

    def info(self, instance):
        if os.path.isfile(parent_path+'/App_Data/'+".Expense_Details.xlsx"):
            df = pd.read_excel(parent_path+'/App_Data/' +
                               ".Expense_Details.xlsx")
            df = str(df)
            self.label.text = df
            self.reset.disabled = False
        else:
            self.label.text = "Please Add Entries!"

    # on text enable add button
    def on_text(self, instance, value):
        a = self.item.text
        b = self.Price.text
        try:
            a = self.item.text
            b = float(self.Price.text)
            if a and b:
                self.add.disabled = False
            else:
                self.add.disabled = True
        except ValueError:
            self.Price.text_hint = "Please enter a valid number!"
            self.add.disabled = True


class MyApp(App):
    def build(self):
        return MyGrid()


if __name__ == '__main__':
    MyApp().run()
